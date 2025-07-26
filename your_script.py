import os
import json
import csv
import threading
from collections import defaultdict
import pandas as pd
from tqdm import tqdm
from tkinter import filedialog, messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# ------------------------ JSON CLEAN PARSER ------------------------ #
def clean_and_parse_json_string(json_string_raw):
    clean_string = json_string_raw.strip()
    first_brace_index = clean_string.find('{')
    if first_brace_index == -1:
        return None
    clean_string = clean_string[first_brace_index:]
    try:
        return json.loads(clean_string)
    except json.JSONDecodeError:
        return None

# ------------------------ DATA EXTRACTION ------------------------ #
def extract_excel_data(json_data, filename):
    extracted_rows = []
    step_counts_for_file = defaultdict(int)
    activity_info = json_data.get('activityinfo', {})
    activity_number = activity_info.get('activityNo', '')
    activity_title = activity_info.get('activityTitle', '')
    reference_id = activity_info.get('referenceID', '')
    levels = ["CORE", "LIGHT-MULTILINGUAL", "MODERATE-MULTILINGUAL", "INTENSIVE-MULTILINGUAL"]
    steps_data = json_data.get('steps', {})
    all_level_steps = {lvl: steps_data.get(lvl, []) for lvl in levels}
    core_steps_list = all_level_steps.get('CORE', [])
    continuous_step_counter = 0
    found_any_steps = False

    for i, core_step in enumerate(core_steps_list):
        found_any_steps = True
        continuous_step_counter += 1
        metadata = core_step.get('metadata', {})
        row = {
            "JSON File name": filename,
            "Reference ID": reference_id,
            "Activity Number": activity_number,
            "Activity Title": activity_title,
            "Number of steps in the core": f"Step {continuous_step_counter}:",
            "Step Title": metadata.get('stepTitle', ''),
            "Name": metadata.get('name', ''),
            "pageReferenceId": str(core_step.get('pageReferenceId', '')),
            "Original Page Sequence": str(core_step.get('originalPageSequence', '')) or ''
        }
        for lvl in levels:
            row[f"{lvl} pageReferenceId"] = ""
        row["CORE pageReferenceId"] = row["pageReferenceId"]
        for lvl in levels:
            if lvl != "CORE" and i < len(all_level_steps[lvl]):
                val = str(all_level_steps[lvl][i].get('pageReferenceId', ''))
                row[f"{lvl} pageReferenceId"] = val if val and val != 'N/A' else ''
        extracted_rows.append(row)
        for lvl in levels:
            step_counts_for_file[lvl] += len(all_level_steps[lvl])
    if not found_any_steps and (activity_number or activity_title or reference_id):
        extracted_rows.append({
            "JSON File name": filename, "Reference ID": reference_id,
            "Activity Number": activity_number, "Activity Title": activity_title,
            "Number of steps in the core": "", "Step Title": "", "Name": "",
            "pageReferenceId": "", "Original Page Sequence": "",
            **{f"{lvl} pageReferenceId": "" for lvl in levels}
        })
    return extracted_rows, step_counts_for_file

def extract_step_number(step_text):
    try:
        return int(str(step_text).strip().split()[1].strip(":"))
    except:
        return None

# ------------------------ MAIN PROCESS ------------------------ #
def process_all(input_folder, output_csv, output_excel, skip_excel=False):
    all_data = []
    step_totals = defaultdict(int)
    ref_tracker = {}
    reused_records = {}

    headers = [
        "JSON File name", "Reference ID", "Activity Number", "Activity Title",
        "Number of steps in the core", "Step Title", "Name", "pageReferenceId",
        "Original Page Sequence", "CORE pageReferenceId",
        "LIGHT-MULTILINGUAL pageReferenceId", "MODERATE-MULTILINGUAL pageReferenceId",
        "INTENSIVE-MULTILINGUAL pageReferenceId"
    ]

    levels = ["CORE", "LIGHT-MULTILINGUAL", "MODERATE-MULTILINGUAL", "INTENSIVE-MULTILINGUAL"]
    modalities = [
        ("CORE pageReferenceId", "CORE Cumulated From", "ADD8E6"),
        ("LIGHT-MULTILINGUAL pageReferenceId", "LIGHT-MULTILINGUAL Cumulated From", "FFD580"),
        ("MODERATE-MULTILINGUAL pageReferenceId", "MODERATE-MULTILINGUAL Cumulated From", "C6EFCE"),
        ("INTENSIVE-MULTILINGUAL pageReferenceId", "INTENSIVE-MULTILINGUAL Cumulated From", "D7BDE2"),
    ]

    for _, col, _ in modalities:
        ref_tracker[col] = {}
        reused_records[col] = {}

    for file in tqdm(os.listdir(input_folder), desc="Processing JSON Files"):
        if file.endswith('.json'):
            with open(os.path.join(input_folder, file), 'r', encoding='utf-8') as f:
                parsed = clean_and_parse_json_string(f.read())
                if parsed:
                    rows, file_counts = extract_excel_data(parsed, file)
                    all_data.extend(rows)
                    for lvl, cnt in file_counts.items():
                        step_totals[lvl] += cnt

    if not all_data:
        messagebox.showerror("No Data", "No valid JSON data found.")
        return

    os.makedirs(os.path.dirname(output_csv), exist_ok=True)
    with open(output_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(all_data)

    if skip_excel:
        messagebox.showinfo("Done", "CSV created. Excel generation skipped.")
        return

    df = pd.read_csv(output_csv)
    df['Activity Number'] = df['Activity Number'].astype(str)
    df['Activity Number Num'] = pd.to_numeric(df['Activity Number'].str.extract(r'(\d+\.?\d*)')[0], errors='coerce')
    df = df.sort_values(by='Activity Number Num', kind='stable')
    df['Step Number'] = df['Number of steps in the core'].apply(extract_step_number)

    for _, cumu_col, _ in modalities:
        df[cumu_col] = "Fresh"

    for idx, row in df.iterrows():
        activity = row['Activity Number']
        step_number = row['Step Number']
        for page_col, cumu_col, _ in modalities:
            page_id = str(row.get(page_col, "")).strip()
            if not page_id:
                continue
            if page_id in ref_tracker[cumu_col]:
                prev_act, prev_step = ref_tracker[cumu_col][page_id]
                ref_str = f"{prev_act} - Step {prev_step}" if prev_act != activity else f"Step {prev_step}"
                df.at[idx, cumu_col] = ref_str
                reused_records[cumu_col].setdefault(page_id, []).append(ref_str)
            ref_tracker[cumu_col][page_id] = (activity, step_number)

    df["Status"] = df.apply(lambda r: "Pass" if len(set(
        r[col] for _, col, _ in modalities if r[col] != "Fresh"
    )) <= 1 else "Fail", axis=1)

    df.drop(columns=['Step Number', 'Activity Number Num'] + [c[0] for c in modalities if "pageReferenceId" in c[0]], inplace=True)
    df.to_excel(output_excel, index=False, sheet_name="Cumulated Data")

    wb = load_workbook(output_excel)
    ws = wb["Cumulated Data"]
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    alt_fills = [PatternFill("solid", fgColor="E6F2FF"), PatternFill("solid", fgColor="FDE6F2")]
    status_colors = {"Pass": PatternFill("solid", fgColor="C6EFCE"), "Fail": PatternFill("solid", fgColor="FFC7CE")}
    fill_fresh = PatternFill("solid", fgColor="F2F2F2")
    fill_map = {col: PatternFill("solid", fgColor=color) for _, col, color in modalities}

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    current_act = None
    toggle = 0
    for row in ws.iter_rows(min_row=2):
        act_val = row[df.columns.get_loc("Activity Number")].value
        status_val = row[df.columns.get_loc("Status")].value
        if act_val != current_act:
            current_act = act_val
            toggle = 1 - toggle
        for i, cell in enumerate(row):
            col_name = df.columns[i]
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            if col_name == "Status":
                cell.fill = status_colors.get(status_val, PatternFill())
            elif col_name in fill_map:
                cell.fill = fill_map[col_name] if str(cell.value).lower() != "fresh" else fill_fresh
            else:
                cell.fill = alt_fills[toggle]

    ws2 = wb.create_sheet("Reused Pages Summary")
    ws2.append(["Modality", "pageReferenceId", "Referenced In"])
    for mod, record in reused_records.items():
        for pid, refs in record.items():
            ws2.append([mod, pid, ", ".join(refs)])
    for col in ws2.columns:
        width = max(len(str(c.value)) if c.value else 0 for c in col)
        ws2.column_dimensions[col[0].column_letter].width = width + 2

    wb.save(output_excel)
    messagebox.showinfo("✅ Done", f"Excel saved at:\n{output_excel}")

# ------------------------ GUI ------------------------ #
import ttkbootstrap as tb
from tkinter import filedialog, messagebox
import threading
import os

def run_gui():
    def browse_input():
        folder = filedialog.askdirectory()
        if folder:
            input_var.set(folder)

    def browse_output():
        file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file:
            output_var.set(file)

    def start_processing():
        if not input_var.get() or not output_var.get():
            messagebox.showwarning("Missing Fields", "Please select both input folder and output Excel file.")
            return
        status_var.set("Processing...")
        csv_path = os.path.splitext(output_var.get())[0] + ".csv"
        threading.Thread(target=lambda: process_wrapper(input_var.get(), csv_path, output_var.get(), skip_var.get())).start()

    def process_wrapper(input_path, csv_path, excel_path, skip_excel):
        try:
            process_all(input_path, csv_path, excel_path, skip_excel)
            status_var.set("✅ Done.")
        except Exception as e:
            status_var.set(f"❌ Error: {str(e)}")

    # UI Setup
    app = tb.Window(themename="cosmo")
    app.title("🧩 Activity Cumulative Report Generator")
    app.geometry("1200x420")
    app.resizable(False, False)
    light_pink = "#ffe6f0"
    app.configure(bg=light_pink)

    input_var = tb.StringVar()
    output_var = tb.StringVar()
    skip_var = tb.BooleanVar()
    status_var = tb.StringVar()

    tb.Label(app, text="Activity Cumulative Report Generator", font=("Segoe UI", 16, "bold"), background=light_pink).pack(pady=15)

    frm = tb.Frame(app, padding=(20, 10), bootstyle="light")
    frm.pack(fill="both", expand=True, padx=30)

    # Input folder
    tb.Label(frm, text="📁 Input Folder:", font=("Segoe UI", 11), background=light_pink).grid(row=0, column=0, sticky="w", pady=10)
    tb.Entry(frm, textvariable=input_var, width=55).grid(row=0, column=1, padx=10)
    tb.Button(frm, text="📂 Browse Folder", bootstyle="warning-gradient", width=18, command=browse_input).grid(row=0, column=2)

    # Output file
    tb.Label(frm, text="📄 Output Excel:", font=("Segoe UI", 11), background=light_pink).grid(row=1, column=0, sticky="w", pady=10)
    tb.Entry(frm, textvariable=output_var, width=55).grid(row=1, column=1, padx=10)
    tb.Button(frm, text="💾 Save As", bootstyle="info-gradient", width=18, command=browse_output).grid(row=1, column=2)

    # Skip Excel checkbox
    tb.Checkbutton(frm, text="Skip Excel Generation", variable=skip_var, bootstyle="round-toggle").grid(row=2, column=1, sticky="w", pady=5)

    # Start button
    tb.Button(app, text="🚀 Start Processing", bootstyle="success-gradient", width=25, command=start_processing).pack(pady=20)

    # Status label
    tb.Label(app, textvariable=status_var, font=("Segoe UI", 10), background=light_pink, foreground="gray").pack(pady=10)

 # Footer with logo and your name
    footer = tb.Frame(app, bootstyle="light", padding=(10, 5))
    footer.pack(side="bottom", fill="x")

    from PIL import Image, ImageTk

    try:
        logo_path = "logo.png"  # Ensure the logo file is in the same folder as your script
        logo_img = Image.open(logo_path)
        logo_img = logo_img.resize((100, 30), Image.Resampling.LANCZOS)
        logo_tk = ImageTk.PhotoImage(logo_img)
        logo_label = tb.Label(footer, image=logo_tk, background="#ffe6f0")
        logo_label.image = logo_tk  # Prevent garbage collection
        logo_label.pack(side="right", padx=10)
    except Exception as e:
        print(f"Logo load failed: {e}")

    # Developer name on the right
    name_label = tb.Label(
        footer,
        text="Tool Created by, HPT - Venkatraman E",
        font=("Segoe UI", 7, "bold"),
        background="#ffe6f0",
        anchor="e"
    )
    name_label.pack(side="right", padx=15)
    app.mainloop()





if __name__ == "__main__":
    run_gui()